from lxml import html
import requests
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import time
#reload(sys)
#sys.setdefaultencoding('utf8')

test1 = 'merverdi'
test2 = 'Merverdi'
test3 = 'MerVerdi'

sites = []
links = []
f=open('regnskapsføreroslo.txt', 'r')
for line in f:
    list = line.split(' ')
    links.append(list[1])

page1 = links[0]
page2 = links[1]



#link = 'https://www.google.com/search?safe=off&ei=hEDwXKnnGNKvrgS03rmgBg&q=regnskapsf%C3%B8rer+oslo&oq=regnskapsf%C3%B8rer+oslo&gs_l=psy-ab.3..0i71l8.0.0..68731...0.0..0.0.0.......0......gws-wiz.F6biEh8buaw'
#link2 = 'https://www.google.com/search?ei=eEDwXKPtMujmrgSG-ICoCg&q=regnskapsf%C3%B8rer+oslo&oq=regnskapsf%C3%B8rer+oslo&gs_l=psy-ab.3..0i71l8.0.0..123644...0.0..0.0.0.......0......gws-wiz.EZPrdvU61Yg'
page = requests.get(page1)
tree = html.fromstring(page.content)

sites.append(tree)#.xpath('//a[@title]/text()'))

first = False
text = page.text.split(' ')
found = 0
placementTest = 0
placement = 0
for word in text:
    if 'href' in word:
        if '/url' in word:
            #print(word)
            placementTest +=1
            if test1 in word:
                found += 1
                first = True
                placement = placementTest
            elif test2 in word:
                found += 1
                first = True
                placement = placementTest
            elif test3 in word:
                found += 1
                first = True
                placement = placementTest

if found == 0:
    #page2 = 'https://www.google.com/search?q=regnskapsf%C3%B8rer+oslo&ei=9UDwXNf4A4bMrgS_p6OgAw&start=10&sa=N&ved=0ahUKEwiXx63Zj8TiAhUGposKHb_TCDQQ8tMDCI8C&biw=1280&bih=610'
    page = requests.get(page2)
    text = page.text.split(' ')
    placementTest = 10
    placement = 0
    for word in text:
        if 'href' in word:
            if '/url' in word:
                #print(word)
                placementTest +=1
                if test1 in word:
                    found += 1
                    placement = placementTest
                elif test2 in word:
                    found += 1
                    placement = placementTest
                elif test3 in word:
                    found += 1
                    placement = placementTest



named_tuple = time.localtime() # get struct_time
time_string = time.strftime("%m.%d.%Y, %H:%M:%S", named_tuple)

wb = load_workbook('google.xlsx')
ws = wb.active
id = ws.cell(row=2, column=1)
current_row = 2
while id.value!= None:
    current_row +=1
    id = ws.cell(row = current_row, column = 1)

if found != 0 and first == True:
    ws.cell(row = current_row, column = 1, value = current_row-1)
    ws.cell(row = current_row, column = 2, value = time_string)
    ws.cell(row = current_row, column = 3, value = placement)
    ws.cell(row = current_row, column = 4, value = 'True')
    ws.cell(row = current_row, column = 5, value = 'MerVerdi')
    ws.cell(row = current_row, column = 6, value = 'regnskapsfører oslo')

elif found != 0 and first == False:
    ws.cell(row = current_row, column = 1, value = current_row-1)
    ws.cell(row = current_row, column = 2, value = time_string)
    ws.cell(row = current_row, column = 3, value = placement)
    ws.cell(row = current_row, column = 4, value = 'False')
    ws.cell(row = current_row, column = 5, value = 'MerVerdi')
    ws.cell(row = current_row, column = 6, value = 'regnskapsfører oslo')

wb.save('google.xlsx')
